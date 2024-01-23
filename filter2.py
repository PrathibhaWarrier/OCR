import requests
import openpyxl

def get_all_data():
    api_url = "https://api.example.com/all_data"
    
    try:
        response = requests.get(api_url)
        response.raise_for_status()  # Check for errors in the response

        all_data = response.json()

        if all_data:
            return all_data
        else:
            print("No data found.")
            return None

    except requests.exceptions.RequestException as e:
        print(f"Error fetching data: {e}")
        return None

def save_to_excel(all_data, output_filename="output.xlsx"):
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Write headers
    headers = ["PartNumber", "TubeType", "TubeDesign", "InnerDiameter", "OuterDiameter"]
    for col_num, header in enumerate(headers, start=1):
        sheet.cell(row=1, column=col_num, value=header)

    # Write data
    row_num = 2
    for data_entry in all_data:
        partnumber = data_entry.get("PartNumber", "")
        sheet.cell(row=row_num, column=1, value=partnumber)  # PartNumber

        for col_num, key in enumerate(["TubeType", "TubeDesign", "InnerDiameter", "OuterDiameter"], start=2):
            sheet.cell(row=row_num, column=col_num, value=data_entry.get(key, ""))

        row_num += 1

    # Save Excel file
    workbook.save(output_filename)
    print(f"Data saved to {output_filename}")

# Example usage:
all_data = get_all_data()

if all_data:
    save_to_excel(all_data)
