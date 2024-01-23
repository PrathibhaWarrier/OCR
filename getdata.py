import requests
import openpyxl

def get_all_data():
    api_url = "https://api.example.com/all_data"
    
    try:
        response = requests.get(api_url)
        response.raise_for_status()  # Check for errors in the response

        all_data = response.json()

        if isinstance(all_data, list) and all_data:
            return all_data
        else:
            print("No valid data found.")
            return None

    except requests.exceptions.RequestException as e:
        print(f"Error fetching data: {e}")
        return None

def save_to_excel(all_data, output_filename="output.xlsx"):
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    if all_data:
        # Write headers dynamically
        headers = list(all_data[0].keys())
        for col_num, header in enumerate(headers, start=1):
            sheet.cell(row=1, column=col_num, value=header)

        # Write data
        row_num = 2
        for data_entry in all_data:
            for col_num, header in enumerate(headers, start=1):
                sheet.cell(row=row_num, column=col_num, value=data_entry.get(header, ""))

            row_num += 1

        # Save Excel file
        workbook.save(output_filename)
        print(f"Data saved to {output_filename}")
    else:
        print("No data to save.")

# Example usage:
all_data = get_all_data()

if all_data is not None:
    save_to_excel(all_data)
